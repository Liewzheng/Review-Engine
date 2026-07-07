from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

def create_test_cases_v2():
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    alt_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    def setup_sheet(ws, title, data):
        ws.title = title
        ws.sheet_view.showGridLines = False
        
        # Header row
        headers = ["用例ID", "分类", "功能", "测试步骤", "测试类型", "优先级", "浏览器", "预期结果", "实际结果"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
        
        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx in [1, 5, 6, 7]:  # ID, type, priority, browser -> center
                    cell.alignment = center_align
                else:
                    cell.alignment = cell_align
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 45
        ws.column_dimensions['I'].width = 12
        
        # Row height for header
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, len(data) + 2):
            ws.row_dimensions[row_idx].height = 60
    
    # ========================================
    # Sheet 1: 全局功能 (Global Features)
    # ========================================
    global_features = [
        ("G-001", "导航", "侧边栏展开", "访问首页，确认侧边栏完整展开，显示所有7个菜单项", "手动", "高", "Chrome/Safari", "侧边栏宽度200px，所有菜单项可见", "待测"),
        ("G-002", "导航", "侧边栏折叠", "点击折叠按钮，确认侧边栏缩至64px，只显示图标", "手动", "高", "Chrome/Safari", "侧边栏收缩，图标居中", "待测"),
        ("G-003", "导航", "移动端侧边栏", "缩小窗口至768px以下，确认侧边栏变为抽屉式", "手动", "中", "Chrome", "侧边栏隐藏，汉堡菜单出现", "待测"),
        ("G-004", "导航", "菜单点击跳转", "依次点击每个菜单项，确认正确跳转到对应页面", "手动", "高", "Chrome/Safari", "URL hash正确变化，页面内容正确", "待测"),
        ("G-005", "导航", "当前菜单高亮", "访问各页面，确认对应菜单项高亮显示", "手动", "高", "Chrome/Safari", "激活项背景色为 --bg-active", "待测"),
        ("G-006", "主题", "深色模式默认", "首次访问，确认页面为深色模式", "手动", "高", "Chrome/Safari", "背景色为 #0f172a", "待测"),
        ("G-007", "主题", "浅色模式切换", "点击主题切换按钮，确认变为浅色模式", "手动", "高", "Chrome/Safari", "背景色变为 #f8fafc", "待测"),
        ("G-008", "主题", "浅色模式持久化", "刷新页面，确认主题设置保持", "手动", "中", "Chrome/Safari", "刷新后仍为上次选择的主题", "待测"),
        ("G-009", "响应式", "桌面布局", "在1280px+窗口，确认4列KPI卡片布局", "手动", "中", "Chrome", "4列网格", "待测"),
        ("G-010", "响应式", "平板布局", "在1024px窗口，确认3列卡片/2列KPI", "手动", "中", "Chrome", "3列或2列网格", "待测"),
        ("G-011", "响应式", "手机布局", "在375px窗口，确认单列布局，无水平滚动", "手动", "中", "Chrome", "单列堆叠，无溢出", "待测"),
        ("G-012", "状态栏", "服务状态显示", "确认顶部状态栏显示 Healthy 绿色徽章（调用 GET /health）", "联调", "高", "Chrome/Safari", "状态徽章为绿色，带脉冲动画，与API返回一致", "待测"),
        ("G-013", "API", "健康检查API", "使用curl调用 GET /health，确认返回200和status", "自动化", "高", "所有", "返回{\"status\":\"healthy\",\"version\":\"x.x.x\"}", "待测"),
        ("G-014", "API", "Metrics端点", "使用curl调用 GET /metrics，确认返回Prometheus格式数据", "自动化", "中", "所有", "返回text/plain格式，包含review_engine_*指标", "待测"),
    ]
    
    ws1 = wb.active
    setup_sheet(ws1, "全局功能", global_features)
    
    # ========================================
    # Sheet 2: Dashboard
    # ========================================
    dashboard = [
        ("D-001", "KPI", "本周审查数", "检查KPI卡片显示从GET /api/v1/dashboard返回的真实reviewCount", "联调", "高", "Chrome/Safari", "显示正确数字（与API一致），趋势箭头向上绿色", "待测"),
        ("D-002", "KPI", "活跃队列", "检查KPI卡片显示从API返回的activeQueue", "联调", "高", "Chrome/Safari", "显示数字与API一致，无趋势", "待测"),
        ("D-003", "KPI", "成功率", "检查KPI卡片显示从API返回的successRate", "联调", "高", "Chrome/Safari", "显示百分比与API一致，趋势正确", "待测"),
        ("D-004", "KPI", "平均耗时", "检查KPI卡片显示从API返回的avgDuration", "联调", "高", "Chrome/Safari", "显示正确时长，趋势正确", "待测"),
        ("D-005", "KPI", "加载骨架", "刷新页面，检查KPI卡片先显示Skeleton加载动画", "手动", "中", "Chrome/Safari", "骨架屏出现，数据加载后平滑过渡消失", "待测"),
        ("D-006", "趋势图", "24h柱状图渲染", "检查24小时活动趋势图用lightweight-charts渲染，确认有24个柱", "手动", "高", "Chrome/Safari", "24个柱子，hover显示tooltip，数据与API一致", "待测"),
        ("D-007", "趋势图", "数据一致性", "对比API返回的trend[]数据与图表显示", "联调", "高", "Chrome/Safari", "图表数据点与API返回完全一致", "待测"),
        ("D-008", "系统健康", "GitLab状态", "检查系统健康栏，GitLab API显示从GET /api/v1/dashboard返回的health.gitlab", "联调", "高", "Chrome/Safari", "绿色点+Connected文字与API状态一致", "待测"),
        ("D-009", "系统健康", "LLM状态", "确认OpenAI/Anthropic/Ollama状态从API health.llm返回", "联调", "高", "Chrome/Safari", "三色状态点与API返回一致", "待测"),
        ("D-010", "系统健康", "整体状态", "确认整体状态由API计算返回，非前端硬编码", "联调", "高", "Chrome/Safari", "状态徽章与API overallStatus一致", "待测"),
        ("D-011", "最近审查", "表格数据", "确认最近审查表格数据来自GET /api/v1/dashboard的recentReviews", "联调", "高", "Chrome/Safari", "表格行数和数据与API一致", "待测"),
        ("D-012", "最近审查", "状态标签", "确认success绿色、failed红色、running蓝色，由API status字段决定", "手动", "高", "Chrome/Safari", "颜色对应API返回的状态值", "待测"),
        ("D-013", "自动刷新", "60秒自动刷新", "等待60秒，确认Dashboard自动调用GET /api/v1/dashboard", "联调", "高", "Chrome/Safari", "网络面板显示新请求，数据更新", "待测"),
        ("D-014", "手动刷新", "刷新按钮", "点击手动刷新按钮，确认重新调用API并更新数据", "联调", "高", "Chrome/Safari", "加载动画出现，数据刷新，Toast提示", "待测"),
        ("D-015", "错误处理", "API 500错误", "模拟后端返回500，确认Dashboard显示错误状态", "联调", "高", "Chrome/Safari", "错误提示出现，显示Retry按钮，不崩溃", "待测"),
        ("D-016", "错误处理", "网络断开", "断开网络，确认显示离线状态提示", "联调", "高", "Chrome/Safari", "网络断开提示，重连后自动恢复", "待测"),
    ]
    
    ws2 = wb.create_sheet()
    setup_sheet(ws2, "Dashboard", dashboard)
    
    # ========================================
    # Sheet 3: Review History
    # ========================================
    review_history = [
        ("H-001", "API查询", "真实搜索", "在搜索框输入关键词，确认调用GET /api/v1/reviews?q=...", "联调", "高", "Chrome/Safari", "请求URL包含q参数，返回结果与前端一致", "待测"),
        ("H-002", "API查询", "项目筛选", "选择项目，确认调用GET /api/v1/reviews?project=...", "联调", "高", "Chrome/Safari", "URL参数正确，返回数据过滤正确", "待测"),
        ("H-003", "API查询", "状态筛选", "选择状态failed，确认调用GET /api/v1/reviews?status=failed", "联调", "高", "Chrome/Safari", "API返回指定状态数据，表格正确显示", "待测"),
        ("H-004", "API查询", "组合筛选", "同时使用搜索+项目+状态筛选，确认URL参数拼接正确", "联调", "高", "Chrome/Safari", "所有参数同时存在于请求URL", "待测"),
        ("H-005", "API查询", "分页参数", "切换分页，确认调用GET /api/v1/reviews?page=2&limit=25", "联调", "高", "Chrome/Safari", "page和limit参数正确，数据正确", "待测"),
        ("H-006", "加载", "骨架屏", "刷新页面，确认先显示骨架屏，等待API响应", "手动", "中", "Chrome/Safari", "el-skeleton显示，加载完成后消失", "待测"),
        ("H-007", "表格", "列排序", "点击表头排序，确认前端排序或请求后端排序", "手动", "中", "Chrome/Safari", "升序/降序切换，数据正确排列", "待测"),
        ("H-008", "表格", "状态列颜色", "确认success绿色、failed红色、running蓝色标签", "手动", "高", "Chrome/Safari", "颜色对应API返回的status值", "待测"),
        ("H-009", "详情", "抽屉数据", "点击某行，确认右侧抽屉显示该行完整数据（API返回的完整对象）", "联调", "高", "Chrome/Safari", "抽屉内容与API返回数据一致，无丢失字段", "待测"),
        ("H-010", "详情", "元信息完整", "确认抽屉显示MR编号、作者、分支、状态、时间、耗时等", "手动", "高", "Chrome/Safari", "所有元信息正确显示，无截断", "待测"),
        ("H-011", "筛选", "重置按钮", "点击重置按钮，确认清除所有筛选并重新请求API", "联调", "高", "Chrome/Safari", "所有筛选器回到默认，表格恢复全部数据", "待测"),
        ("H-012", "API", "空结果处理", "输入不存在的关键词，确认API返回空数组，前端显示空状态", "联调", "中", "Chrome/Safari", "el-empty显示，无错误提示", "待测"),
        ("H-013", "API", "大数据量分页", "确认总页数由API返回的totalCount计算", "联调", "高", "Chrome/Safari", "分页器页数与totalCount/limit一致", "待测"),
        ("H-014", "错误处理", "API超时", "模拟API延迟10秒，确认超时处理和错误提示", "联调", "中", "Chrome/Safari", "超时提示，显示重试按钮", "待测"),
        ("H-015", "错误处理", "网络错误", "断开网络后搜索，确认错误状态处理", "联调", "高", "Chrome/Safari", "网络错误提示，恢复后自动重试", "待测"),
    ]
    
    ws3 = wb.create_sheet()
    setup_sheet(ws3, "Review History", review_history)
    
    # ========================================
    # Sheet 4: Configuration
    # ========================================
    config = [
        ("C-001", "API加载", "配置加载", "刷新页面，确认调用GET /api/v1/config加载当前配置", "联调", "高", "Chrome/Safari", "API返回200，表单字段正确填充", "待测"),
        ("C-002", "API加载", "数据一致性", "对比API返回的JSON与前端表单显示值", "联调", "高", "Chrome/Safari", "所有字段值与API返回一致，无错位", "待测"),
        ("C-003", "模式", "只读模式", "默认进入，确认所有字段不可编辑（Edit按钮可见）", "手动", "高", "Chrome/Safari", "表单禁用，Edit按钮可见", "待测"),
        ("C-004", "模式", "编辑模式", "点击Edit按钮，确认字段变为可编辑", "手动", "高", "Chrome/Safari", "输入框可编辑，Save/Cancel按钮出现", "待测"),
        ("C-005", "模式", "取消编辑", "点击Cancel，确认恢复只读模式，不调用API", "手动", "高", "Chrome/Safari", "表单恢复禁用，修改不保存，不发送请求", "待测"),
        ("C-006", "敏感字段", "API Token掩码", "只读模式下，确认API Token显示为掩码", "手动", "高", "Chrome/Safari", "掩码显示，眼睛图标可点击", "待测"),
        ("C-007", "敏感字段", "Token显示", "点击眼睛图标，确认Token显示5秒", "手动", "高", "Chrome/Safari", "Token明文显示，5秒后自动隐藏", "待测"),
        ("C-008", "API保存", "PUT保存", "修改后点击Save，确认调用PUT /api/v1/config（完整配置）", "联调", "高", "Chrome/Safari", "请求体为完整配置JSON，返回200", "待测"),
        ("C-009", "API保存", "保存反馈", "保存成功后，确认Toast提示和状态同步到AppConfig", "联调", "高", "Chrome/Safari", "Toast成功通知，其他页面配置同步更新", "待测"),
        ("C-010", "API测试", "测试连接", "点击Test Connection，确认调用POST /api/v1/config/test", "联调", "高", "Chrome/Safari", "返回latencyMs，显示Toast结果", "待测"),
        ("C-011", "API测试", "测试失败", "模拟错误Token，确认测试返回错误，latency为null", "联调", "高", "Chrome/Safari", "错误提示，latency显示N/A", "待测"),
        ("C-012", "表单", "LLM Provider联动", "选择不同provider，确认模型选项变化", "手动", "高", "Chrome/Safari", "模型下拉框内容变化", "待测"),
        ("C-013", "表单", "Temperature滑块", "拖动滑块，确认数值实时变化", "手动", "中", "Chrome/Safari", "滑块移动，数值同步更新", "待测"),
        ("C-014", "表单", "未保存警告", "修改后尝试离开页面，确认浏览器确认对话框", "手动", "高", "Chrome/Safari", "浏览器确认对话框弹出，阻止导航", "待测"),
        ("C-015", "错误处理", "PUT 500错误", "模拟PUT返回500，确认保存失败提示", "联调", "高", "Chrome/Safari", "错误Toast，保持编辑模式，不丢失修改", "待测"),
        ("C-016", "加载", "骨架屏", "刷新页面，确认先显示el-skeleton，等待GET /api/v1/config响应", "手动", "中", "Chrome/Safari", "骨架屏显示，加载后平滑过渡", "待测"),
    ]
    
    ws4 = wb.create_sheet()
    setup_sheet(ws4, "Configuration", config)
    
    # ========================================
    # Sheet 5: Queue Monitor
    # ========================================
    queue = [
        ("Q-001", "API加载", "队列统计", "刷新页面，确认调用GET /api/v1/queue/stats", "联调", "高", "Chrome/Safari", "API返回200，统计数字正确显示", "待测"),
        ("Q-002", "API数据", "isPaused计算", "确认前端isPaused计算属性绑定到stats.isPaused", "联调", "高", "Chrome/Safari", "isPaused与API返回值一致，UI正确响应", "待测"),
        ("Q-003", "API数据", "队列深度计算", "确认队列深度 = active + queued（由API返回或前端计算）", "手动", "高", "Chrome/Safari", "计算正确，与API数据一致", "待测"),
        ("Q-004", "任务卡片", "状态显示", "确认running任务有进度条和专家名（API返回数据）", "手动", "高", "Chrome/Safari", "卡片内容完整，数据与API一致", "待测"),
        ("Q-005", "API控制", "暂停队列", "点击Pause，确认调用POST /api/v1/queue/pause", "联调", "高", "Chrome/Safari", "返回200，按钮变Resume，queued显示暂停", "待测"),
        ("Q-006", "API控制", "恢复队列", "点击Resume，确认调用POST /api/v1/queue/resume", "联调", "高", "Chrome/Safari", "返回200，按钮变Pause，任务继续处理", "待测"),
        ("Q-007", "API控制", "调整并发", "调整最大并发数，确认调用POST /api/v1/queue/max-concurrent", "联调", "高", "Chrome/Safari", "请求体包含maxConcurrent值，返回200", "待测"),
        ("Q-008", "API控制", "取消任务", "点击Cancel，确认调用DELETE /api/v1/queue/tasks/{id}", "联调", "高", "Chrome/Safari", "确认对话框后发送DELETE，任务移除", "待测"),
        ("Q-009", "实时更新", "轮询刷新", "等待自动刷新间隔，确认重新调用GET /api/v1/queue/stats", "联调", "高", "Chrome/Safari", "数据自动更新，进度条变化", "待测"),
        ("Q-010", "统计", "进度条", "确认进度条显示正确百分比（由API progress字段）", "手动", "中", "Chrome/Safari", "颜色对应状态，数值与API一致", "待测"),
        ("Q-011", "空状态", "无任务", "清空所有任务，确认显示el-empty", "手动", "中", "Chrome/Safari", "空状态图标和文字", "待测"),
        ("Q-012", "加载", "骨架屏", "刷新页面，确认先显示骨架屏，等待API响应", "手动", "中", "Chrome/Safari", "el-skeleton显示，加载后消失", "待测"),
        ("Q-013", "错误处理", "暂停失败", "模拟POST /queue/pause返回500，确认错误提示", "联调", "中", "Chrome/Safari", "错误Toast，按钮状态不切换", "待测"),
        ("Q-014", "错误处理", "取消失败", "模拟DELETE返回404，确认任务不存在提示", "联调", "中", "Chrome/Safari", "错误提示，任务从列表移除（已不存在）", "待测"),
    ]
    
    ws5 = wb.create_sheet()
    setup_sheet(ws5, "Queue Monitor", queue)
    
    # ========================================
    # Sheet 6: LLM Status
    # ========================================
    llm = [
        ("L-001", "API加载", "提供商列表", "刷新页面，确认调用GET /api/v1/llm/providers", "联调", "高", "Chrome/Safari", "API返回200，所有provider卡片正确渲染", "待测"),
        ("L-002", "API数据", "提供商信息", "确认每个provider显示名称、状态、延迟（来自API）", "联调", "高", "Chrome/Safari", "卡片内容与API返回数据一致", "待测"),
        ("L-003", "API数据", "状态颜色", "确认绿色=success、黄色=warning、红色=error（由API status字段决定）", "联调", "高", "Chrome/Safari", "颜色对应API状态值正确", "待测"),
        ("L-004", "API数据", "延迟颜色", "确认延迟>1s为黄色、>2s为红色（由API latencyMs决定）", "联调", "中", "Chrome/Safari", "颜色编码与API返回值一致", "待测"),
        ("L-005", "API数据", "sparkline", "确认24h延迟趋势图数据来自API latencyHistory", "手动", "中", "Chrome/Safari", "SVG折线图数据与API一致", "待测"),
        ("L-006", "API测试", "单provider测试", "点击Test按钮，确认调用POST /api/v1/llm/providers/{id}/test", "联调", "高", "Chrome/Safari", "按钮loading，返回latencyMs，结果显示", "待测"),
        ("L-007", "API测试", "测试失败", "模拟错误配置，确认测试返回错误", "联调", "高", "Chrome/Safari", "错误提示，latency显示N/A", "待测"),
        ("L-008", "批量刷新", "Refresh All", "点击Refresh All，确认依次调用所有provider测试API", "联调", "中", "Chrome/Safari", "所有卡片loading，Toast汇总结果", "待测"),
        ("L-009", "统计", "汇总行", "确认顶部汇总卡片由API数据计算（健康/降级/错误/离线数量）", "联调", "中", "Chrome/Safari", "数字与API返回的providers统计一致", "待测"),
        ("L-010", "配置跳转", "跳转配置", "点击Configure，确认跳转至Config页面", "手动", "中", "Chrome/Safari", "路由跳转至/#/config", "待测"),
        ("L-011", "加载", "骨架屏", "刷新页面，确认先显示el-skeleton，等待API响应", "手动", "中", "Chrome/Safari", "骨架屏显示，加载后消失", "待测"),
        ("L-012", "错误处理", "API错误", "模拟GET /llm/providers返回500，确认错误状态", "联调", "高", "Chrome/Safari", "错误提示，显示Retry按钮", "待测"),
        ("L-013", "实时更新", "自动刷新", "等待自动刷新周期，确认重新调用GET /llm/providers", "联调", "中", "Chrome/Safari", "网络面板显示新请求，数据更新", "待测"),
    ]
    
    ws6 = wb.create_sheet()
    setup_sheet(ws6, "LLM Status", llm)
    
    # ========================================
    # Sheet 7: System Logs
    # ========================================
    logs = [
        ("LG-001", "SSE连接", "实时日志流", "刷新页面，确认建立GET /api/v1/logs SSE连接", "联调", "高", "Chrome/Safari", "EventSource连接建立，返回text/event-stream", "待测"),
        ("LG-002", "SSE数据", "日志格式", "确认SSE返回的日志格式为NDJSON，前端正确解析", "联调", "高", "Chrome/Safari", "每行日志正确解析level、message、timestamp", "待测"),
        ("LG-003", "显示", "终端风格", "确认日志背景为黑色，字体等宽（终端模拟风格）", "手动", "高", "Chrome/Safari", "黑色背景，monospace字体", "待测"),
        ("LG-004", "显示", "级别颜色", "确认INFO白色、WARN黄色、ERROR红色、DEBUG灰色（由API level字段）", "联调", "高", "Chrome/Safari", "颜色对应API返回的level值", "待测"),
        ("LG-005", "显示", "时间戳", "确认时间戳格式正确（由API timestamp字段）", "联调", "中", "Chrome/Safari", "时间戳可读，格式一致", "待测"),
        ("LG-006", "过滤", "级别筛选", "取消勾选INFO，确认前端过滤不显示INFO日志", "手动", "高", "Chrome/Safari", "INFO日志消失，其他级别保留", "待测"),
        ("LG-007", "过滤", "关键词搜索", "输入关键词，确认前端过滤日志行", "手动", "高", "Chrome/Safari", "只显示匹配的日志行", "待测"),
        ("LG-008", "控制", "暂停接收", "点击Pause，确认关闭EventSource连接", "联调", "高", "Chrome/Safari", "EventSource关闭，新日志不显示，计数增加", "待测"),
        ("LG-009", "控制", "恢复接收", "点击Resume，确认重新建立EventSource连接", "联调", "高", "Chrome/Safari", "EventSource重新连接，日志继续滚动，计数清零", "待测"),
        ("LG-010", "API下载", "NDJSON下载", "点击Download，确认调用GET /api/v1/logs/download", "联调", "中", "Chrome/Safari", "返回NDJSON文件，自动下载，Toast通知", "待测"),
        ("LG-011", "控制", "清除", "点击Clear，确认前端日志清空（不调用API）", "手动", "中", "Chrome/Safari", "日志清空，确认对话框，恢复后SSE继续", "待测"),
        ("LG-012", "实时", "新日志提示", "暂停时等待新日志，确认浮动按钮显示新日志数量", "手动", "中", "Chrome/Safari", "浮动按钮显示新日志数量，点击可查看", "待测"),
        ("LG-013", "实时", "自动滚动", "确认日志自动滚动到底部（SSE收到新消息时）", "手动", "高", "Chrome/Safari", "最新日志始终在底部可见", "待测"),
        ("LG-014", "持久化", "日志文件", "确认后端日志持久化到~/.config/review-engine/logs.ndjson", "自动化", "中", "所有", "文件存在，包含SSE发送的相同日志内容", "待测"),
        ("LG-015", "错误处理", "SSE断开", "模拟SSE断开，确认自动重连机制", "联调", "高", "Chrome/Safari", "EventSource自动重连，日志恢复接收", "待测"),
        ("LG-016", "响应式", "移动端", "375px窗口，确认字体缩小，按钮堆叠", "手动", "中", "Chrome", "布局适配，日志可读", "待测"),
    ]
    
    ws7 = wb.create_sheet()
    setup_sheet(ws7, "System Logs", logs)
    
    # ========================================
    # Sheet 8: Experts Management
    # ========================================
    experts = [
        ("E-001", "API加载", "专家列表", "刷新页面，确认调用GET /api/v1/system/experts", "联调", "高", "Chrome/Safari", "API返回200，所有专家卡片正确渲染", "待测"),
        ("E-002", "API数据", "卡片显示", "确认每个卡片显示图标、名称、类别、权重、描述（来自API）", "联调", "高", "Chrome/Safari", "卡片内容与API返回数据一致", "待测"),
        ("E-003", "API数据", "类别颜色", "确认不同类别有不同颜色标签（由API category字段决定）", "手动", "中", "Chrome/Safari", "颜色区分不同类别，与API一致", "待测"),
        ("E-004", "API更新", "启用禁用", "点击toggle开关，确认调用PUT /api/v1/system/experts/{id}", "联调", "高", "Chrome/Safari", "请求体包含enabled字段，返回200，Toast通知", "待测"),
        ("E-005", "API更新", "权重更新", "拖动权重滑块保存，确认调用PUT /api/v1/system/experts/{id}", "联调", "高", "Chrome/Safari", "请求体包含weight字段，返回200，卡片更新", "待测"),
        ("E-006", "API一致性", "PUT往返验证", "修改enabled/weight后刷新，确认GET返回的值与PUT一致", "联调", "高", "Chrome/Safari", "GET返回的值与PUT发送的值一致（持久化验证）", "待测"),
        ("E-007", "筛选", "搜索", "输入专家名称，确认前端过滤卡片", "手动", "高", "Chrome/Safari", "只显示匹配卡片", "待测"),
        ("E-008", "筛选", "类别筛选", "选择类别，确认只显示该类别", "手动", "中", "Chrome/Safari", "卡片按类别过滤", "待测"),
        ("E-009", "详情", "弹窗数据", "点击View Details，确认弹窗显示该专家完整API数据", "联调", "高", "Chrome/Safari", "ElDialog显示，内容与API返回一致", "待测"),
        ("E-010", "详情", "提示词预览", "确认弹窗显示专家提示词（monospace，来自API prompt字段）", "手动", "中", "Chrome/Safari", "提示词正确显示，可复制", "待测"),
        ("E-011", "详情", "最近审查", "确认弹窗显示最近审查记录（来自API recentReviews）", "联调", "中", "Chrome/Safari", "表格数据与API一致，可点击跳转", "待测"),
        ("E-012", "编辑", "编辑弹窗", "点击Edit按钮，确认编辑弹窗打开，表单预填充API数据", "手动", "高", "Chrome/Safari", "表单可编辑，初始值与API一致", "待测"),
        ("E-013", "编辑", "权重滑块", "拖动权重滑块，确认数值实时变化", "手动", "中", "Chrome/Safari", "滑块移动，数值更新", "待测"),
        ("E-014", "统计", "顶部统计", "确认顶部active数和平均权重由API数据计算", "联调", "中", "Chrome/Safari", "数字与API返回的专家列表统计一致", "待测"),
        ("E-015", "响应式", "网格布局", "确认1280px 4列、1024px 3列、768px 2列、375px 1列", "手动", "中", "Chrome", "断点正确，布局适配", "待测"),
        ("E-016", "错误处理", "PUT失败", "模拟PUT返回500，确认错误提示，状态不切换", "联调", "高", "Chrome/Safari", "错误Toast，开关/滑块恢复原始状态", "待测"),
    ]
    
    ws8 = wb.create_sheet()
    setup_sheet(ws8, "Experts Management", experts)
    
    # ========================================
    # Sheet 9: API Integration (NEW)
    # ========================================
    api_integration = [
        ("A-001", "数据一致性", "Dashboard GET/PUT往返", "调用GET /api/v1/dashboard获取数据，验证前端显示与API一致；修改后端数据后刷新验证", "联调", "高", "Chrome/Safari", "前端显示值与API原始返回值完全一致", "待测"),
        ("A-002", "数据一致性", "Config GET/PUT往返", "调用GET /api/v1/config获取配置，修改后PUT保存，再次GET验证持久化", "联调", "高", "Chrome/Safari", "PUT后GET返回的值与PUT请求体一致", "待测"),
        ("A-003", "数据一致性", "Experts PUT往返", "修改专家enabled/weight后，调用GET /api/v1/system/experts/{id}验证", "联调", "高", "Chrome/Safari", "修改后的值在GET中正确返回", "待测"),
        ("A-004", "数据一致性", "Queue stats实时性", "调用GET /api/v1/queue/stats，验证前端统计与API返回一致", "联调", "高", "Chrome/Safari", "active/queued/paused数值与API一致", "待测"),
        ("A-005", "数据一致性", "Reviews分页一致性", "调用GET /api/v1/reviews?page=N&limit=M，验证前端表格行数与API返回一致", "联调", "高", "Chrome/Safari", "表格数据与API data[]一致，totalCount与分页器一致", "待测"),
        ("A-006", "错误处理", "网络断开", "断开容器网络或停止后端，确认前端优雅降级", "联调", "高", "Chrome/Safari", "错误提示出现，不白屏，重连后自动恢复", "待测"),
        ("A-007", "错误处理", "500错误处理", "模拟后端返回500，确认各页面显示错误状态", "联调", "高", "Chrome/Safari", "错误Toast/提示，显示Retry按钮，可重试", "待测"),
        ("A-008", "错误处理", "超时处理", "模拟API延迟10s+，确认前端超时处理", "联调", "高", "Chrome/Safari", "超时提示，取消请求，不阻塞UI", "待测"),
        ("A-009", "错误处理", "404处理", "请求不存在的API端点，确认404处理", "自动化", "中", "所有", "返回404，前端不崩溃", "待测"),
        ("A-010", "加载状态", "Skeleton加载", "刷新各页面，确认先显示Skeleton加载动画，再显示数据", "手动", "高", "Chrome/Safari", "Skeleton显示，数据加载后平滑过渡", "待测"),
        ("A-011", "加载状态", "加载超时", "模拟慢速API，确认Skeleton显示超过3秒", "联调", "中", "Chrome/Safari", "Skeleton持续显示直到数据加载完成或超时", "待测"),
        ("A-012", "加载状态", "加载取消", "快速切换页面，确认未完成的请求被取消（AbortController）", "联调", "中", "Chrome/Safari", "网络面板显示请求已取消，无内存泄漏", "待测"),
        ("A-013", "实时更新", "SSE连接", "确认System Logs页面建立持久SSE连接，接收实时数据", "联调", "高", "Chrome/Safari", "EventSource连接保持，实时接收日志", "待测"),
        ("A-014", "实时更新", "SSE重连", "断开SSE连接，确认自动重连机制", "联调", "高", "Chrome/Safari", "EventSource自动重连，恢复日志接收", "待测"),
        ("A-015", "实时更新", "Dashboard轮询", "确认Dashboard每60秒自动调用GET /api/v1/dashboard", "联调", "高", "Chrome/Safari", "网络面板显示定时请求，数据自动更新", "待测"),
        ("A-016", "实时更新", "Queue轮询", "确认Queue Monitor自动刷新stats数据", "联调", "高", "Chrome/Safari", "定时请求GET /api/v1/queue/stats，数据更新", "待测"),
        ("A-017", "API验证", "curl验证后端", "使用curl逐一验证所有可用API端点", "自动化", "高", "所有", "所有端点返回正确HTTP状态和JSON格式", "待测"),
        ("A-018", "API验证", "CORS验证", "确认前端跨域请求正常，响应头包含Access-Control-Allow-Origin", "自动化", "中", "所有", "CORS头正确，预检请求通过", "待测"),
        ("A-019", "API验证", "Content-Type", "确认所有API响应Content-Type为application/json", "自动化", "中", "所有", "响应头Content-Type正确", "待测"),
        ("A-020", "API验证", "API版本一致性", "确认所有API使用/v1前缀，版本一致", "自动化", "低", "所有", "所有请求路径包含/api/v1/", "待测"),
    ]
    
    ws9 = wb.create_sheet()
    setup_sheet(ws9, "API Integration", api_integration)
    
    # Save
    output_path = "/Users/isletspace/Workspace/gitlab.islet.space/review-engine/frontend-test-cases-v2.xlsx"
    wb.save(output_path)
    
    # Verify
    file_size = os.path.getsize(output_path)
    
    result = {
        "output_path": output_path,
        "file_size_bytes": file_size,
        "file_size_kb": round(file_size / 1024, 2),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "test_cases_per_sheet": {
            name: ws.max_row - 1 for name, ws in zip(wb.sheetnames, wb.worksheets)
        }
    }
    
    print(result)
    return result

create_test_cases_v2()

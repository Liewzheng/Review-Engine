import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_v2_workbook():
    wb = openpyxl.Workbook()
    
    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 列标题
    headers = ["用例ID", "分类", "功能", "测试步骤", "测试类型", "优先级", "浏览器", "预期结果", "实际结果"]
    
    # ===== 1. 全局功能 =====
    ws = wb.active
    ws.title = "全局功能"
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    global_cases = [
        ["G-001", "导航", "侧边栏展开", "访问首页，确认侧边栏完整展开，显示所有7个菜单项", "手动", "高", "Chrome/Safari", "侧边栏宽度200px，所有菜单项可见", "PASS/待测"],
        ["G-002", "导航", "侧边栏折叠", "点击折叠按钮，确认侧边栏缩至64px，只显示图标", "手动", "高", "Chrome/Safari", "侧边栏收缩，图标居中", "PASS/待测"],
        ["G-003", "导航", "移动端侧边栏", "缩小窗口至768px以下，确认侧边栏变为抽屉式", "手动", "中", "Chrome", "侧边栏隐藏，汉堡菜单出现", "PASS/待测"],
        ["G-004", "导航", "菜单点击跳转", "依次点击每个菜单项，确认正确跳转到对应页面", "手动", "高", "Chrome/Safari", "URL hash正确变化，页面内容正确", "PASS/待测"],
        ["G-005", "导航", "当前菜单高亮", "访问各页面，确认对应菜单项高亮显示", "手动", "高", "Chrome/Safari", "激活项背景色为 --bg-active", "PASS/待测"],
        ["G-006", "主题", "深色模式默认", "首次访问，确认页面为深色模式", "手动", "高", "Chrome/Safari", "背景色为 #0f172a", "PASS/待测"],
        ["G-007", "主题", "浅色模式切换", "点击主题切换按钮，确认变为浅色模式", "手动", "高", "Chrome/Safari", "背景色变为 #f8fafc", "PASS/待测"],
        ["G-008", "主题", "浅色模式持久化", "刷新页面，确认主题设置保持", "手动", "中", "Chrome/Safari", "刷新后仍为上次选择的主题", "PASS/待测"],
        ["G-009", "响应式", "桌面布局", "在1280px+窗口，确认4列KPI卡片布局", "手动", "中", "Chrome", "4列网格", "PASS/待测"],
        ["G-010", "响应式", "平板布局", "在1024px窗口，确认3列卡片/2列KPI", "手动", "中", "Chrome", "3列或2列网格", "PASS/待测"],
        ["G-011", "响应式", "手机布局", "在375px窗口，确认单列布局，无水平滚动", "手动", "中", "Chrome", "单列堆叠，无溢出", "PASS/待测"],
        ["G-012", "状态栏", "服务状态显示", "确认顶部状态栏显示 Healthy 绿色徽章", "手动", "高", "Chrome/Safari", "状态徽章为绿色，带脉冲动画", "PASS/待测"],
        ["G-013", "API联调", "健康检查API", "curl GET /api/v1/health，确认返回200及healthy状态", "自动", "高", "所有", "JSON返回{status: healthy}", "待测"],
        ["G-014", "API联调", "页面加载性能", "使用Performance API测量首屏加载时间", "自动", "中", "Chrome", "首屏加载<2s，API响应<1s", "待测"],
        ["G-015", "API联调", "404页面处理", "访问不存在的路由如/#/nonexistent，确认显示404", "手动", "中", "Chrome/Safari", "显示404页面或重定向至首页", "待测"],
        ["G-016", "API联调", "浏览器兼容性", "在Firefox中验证导航和主题切换功能", "手动", "低", "Firefox", "功能正常，布局无错乱", "待测"],
    ]
    for row in global_cases:
        ws.append(row)
    
    # ===== 2. Dashboard =====
    ws = wb.create_sheet("Dashboard")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    dashboard_cases = [
        ["D-001", "KPI", "本周审查数", "检查KPI卡片显示数字，趋势百分比正确", "手动", "高", "Chrome/Safari", "显示正确数字，趋势箭头向上绿色", "PASS/待测"],
        ["D-002", "KPI", "活跃队列", "检查KPI卡片显示active队列数量", "手动", "高", "Chrome/Safari", "显示正确数值，无趋势时显示—", "PASS/待测"],
        ["D-003", "KPI", "成功率", "检查KPI卡片显示成功率百分比", "手动", "高", "Chrome/Safari", "显示正确百分比，趋势箭头正确", "PASS/待测"],
        ["D-004", "KPI", "平均耗时", "检查KPI卡片显示平均审查时长", "手动", "高", "Chrome/Safari", "显示正确时长，趋势箭头正确", "PASS/待测"],
        ["D-005", "KPI", "加载骨架", "刷新页面，检查KPI卡片先显示骨架屏", "手动", "中", "Chrome/Safari", "骨架屏出现，数据加载后消失", "PASS/待测"],
        ["D-006", "趋势图", "24h柱状图", "检查24小时活动趋势图，确认有24个柱子", "手动", "高", "Chrome/Safari", "24个柱子，hover显示tooltip", "PASS/待测"],
        ["D-007", "趋势图", "工作日峰值", "确认9-18时柱子明显高于其他时段", "手动", "中", "Chrome/Safari", "工作时间柱子更高", "PASS/待测"],
        ["D-008", "系统健康", "GitLab状态", "检查系统健康栏，GitLab API显示状态正确", "手动", "高", "Chrome/Safari", "绿色点+Operational/灰色点+Offline", "PASS/待测"],
        ["D-009", "系统健康", "GitHub状态", "确认GitHub API显示状态正确", "手动", "中", "Chrome/Safari", "状态点颜色+文字正确", "PASS/待测"],
        ["D-010", "系统健康", "LLM状态", "确认OpenAI/Anthropic/Ollama状态显示正确", "手动", "高", "Chrome/Safari", "三色状态点正确", "PASS/待测"],
        ["D-011", "系统健康", "整体状态", "确认整体状态计算正确", "手动", "高", "Chrome/Safari", "徽章颜色和文字与最差的子系统一致", "PASS/待测"],
        ["D-012", "最近审查", "表格列", "确认表格有MR标题、作者、状态、耗时、时间列", "手动", "高", "Chrome/Safari", "5列正确显示", "PASS/待测"],
        ["D-013", "最近审查", "状态标签", "确认success绿色、failed红色、running蓝色、queued灰色", "手动", "高", "Chrome/Safari", "颜色对应正确", "PASS/待测"],
        ["D-014", "最近审查", "行点击", "点击某行，确认跳转至History详情", "手动", "中", "Chrome/Safari", "路由跳转至/#/history?reviewId=xxx", "PASS/待测"],
        ["D-015", "自动刷新", "定时刷新", "等待60秒，确认数据自动刷新", "手动", "低", "Chrome", "Last Updated时间更新，数据变化", "PASS/待测"],
        ["D-016", "手动刷新", "刷新按钮", "点击刷新按钮，确认数据更新", "手动", "中", "Chrome/Safari", "数据更新，Toast通知出现", "PASS/待测"],
        ["D-017", "API联调", "GET /api/v1/dashboard 响应格式", "curl 调用API，确认返回kpi/trend/health/recentReviews结构", "自动", "高", "所有", "JSON包含4个字段，数据类型正确", "待测"],
        ["D-018", "API联调", "lightweight-charts 渲染正确", "检查趋势图DOM，确认canvas渲染24个数据点", "手动", "高", "Chrome", "lightweight-charts canvas存在，24个点", "待测"],
        ["D-019", "API联调", "60s自动刷新数据一致性", "等待自动刷新后，curl API对比前后数据，确认前端与后端一致", "自动", "高", "Chrome", "前端显示数据与API返回一致", "待测"],
        ["D-020", "API联调", "手动刷新 API 数据往返验证", "点击刷新按钮后，curl API获取最新数据，与前端显示对比", "自动", "高", "Chrome", "前端显示数据与API最新返回一致", "待测"],
        ["D-021", "API联调", "后端500错误时前端错误处理", "模拟后端返回500，确认前端显示错误提示而非空白", "手动", "高", "Chrome", "显示错误状态/Toast提示，保留上次数据", "待测"],
        ["D-022", "API联调", "网络断开时 Skeleton 加载", "断网后刷新页面，确认Skeleton持续显示，网络恢复后数据加载", "手动", "高", "Chrome", "Skeleton持续显示，网络恢复后数据刷新", "待测"],
        ["D-023", "API联调", "空数据状态", "首次使用无审查数据时，确认KPI显示0/—，趋势图为空", "手动", "中", "Chrome", "KPI显示0或—，趋势图显示空状态", "待测"],
        ["D-024", "API联调", "数据更新时间戳显示", "确认Last Updated时间戳与API响应时间一致", "自动", "中", "Chrome", "时间戳格式正确，与API响应时间偏差<1s", "待测"],
    ]
    for row in dashboard_cases:
        ws.append(row)
    
    # ===== 3. Review History =====
    ws = wb.create_sheet("Review History")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    history_cases = [
        ["H-001", "筛选", "搜索框", "在搜索框输入关键词，确认结果过滤", "手动", "高", "Chrome/Safari", "只显示包含关键词的记录", "PASS/待测"],
        ["H-002", "筛选", "项目筛选", "选择项目，确认只显示对应项目记录", "手动", "高", "Chrome/Safari", "表格按项目过滤", "PASS/待测"],
        ["H-003", "筛选", "状态筛选", "选择状态，确认只显示对应状态记录", "手动", "高", "Chrome/Safari", "只显示对应状态", "PASS/待测"],
        ["H-004", "筛选", "日期范围", "选择日期范围，确认结果在范围内", "手动", "中", "Chrome/Safari", "只显示日期范围内的记录", "PASS/待测"],
        ["H-005", "筛选", "仓库筛选", "选择仓库筛选，确认过滤正确", "手动", "中", "Chrome/Safari", "按仓库过滤", "PASS/待测"],
        ["H-006", "筛选", "重置按钮", "点击重置按钮，确认所有筛选器清空", "手动", "高", "Chrome/Safari", "所有筛选器回到默认，表格恢复全部数据", "PASS/待测"],
        ["H-007", "分页", "每页25条", "选择每页25条，确认显示25行", "手动", "高", "Chrome/Safari", "表格显示25行，分页器正确", "PASS/待测"],
        ["H-008", "分页", "每页50条", "选择每页50条，确认显示50行", "手动", "高", "Chrome/Safari", "表格显示50行", "PASS/待测"],
        ["H-009", "分页", "每页100条", "选择每页100条，确认显示100行", "手动", "高", "Chrome/Safari", "表格显示100行", "PASS/待测"],
        ["H-010", "分页", "翻页", "点击下一页，确认数据变化，URL更新", "手动", "高", "Chrome/Safari", "URL page参数变化，数据更新", "PASS/待测"],
        ["H-011", "表格", "列排序", "点击表头排序，确认数据重新排序", "手动", "中", "Chrome/Safari", "升序/降序切换", "PASS/待测"],
        ["H-012", "表格", "状态列颜色", "确认success绿色、failed红色、running蓝色标签", "手动", "高", "Chrome/Safari", "颜色对应正确", "PASS/待测"],
        ["H-013", "详情", "抽屉打开", "点击某行，确认右侧详情抽屉打开", "手动", "高", "Chrome/Safari", "抽屉从右侧滑入，显示详情", "PASS/待测"],
        ["H-014", "详情", "元信息", "确认抽屉显示MR编号、作者、分支、状态等", "手动", "高", "Chrome/Safari", "所有元信息正确显示", "PASS/待测"],
        ["H-015", "详情", "专家结果", "确认抽屉显示各专家审查结果", "手动", "中", "Chrome/Safari", "专家列表和评分显示", "PASS/待测"],
        ["H-016", "详情", "重新审查", "点击重新审查按钮，确认有反馈", "手动", "中", "Chrome/Safari", "Toast通知或确认对话框", "PASS/待测"],
        ["H-017", "空状态", "无结果", "搜索不存在的词，确认显示el-empty", "手动", "中", "Chrome/Safari", "显示空状态图标和文字", "PASS/待测"],
        ["H-018", "API联调", "GET /api/v1/reviews 后端筛选验证", "curl 调用API带status参数，确认返回结果与前端筛选一致", "自动", "高", "所有", "API返回status过滤后的数据，前端显示一致", "待测"],
        ["H-019", "API联调", "分页参数 URL 同步", "切换分页后，curl 对应page/size参数，确认后端返回正确分页", "自动", "高", "Chrome", "URL page/size与API参数一致，数据正确", "待测"],
        ["H-020", "API联调", "后端搜索 q 参数正确传递", "在搜索框输入关键词，通过Network面板确认q参数传递至API", "手动", "高", "Chrome", "Network中请求URL包含?q=关键词", "待测"],
        ["H-021", "API联调", "空搜索结果 API 返回", "搜索不存在的关键词，curl API确认返回空数组，前端显示空状态", "自动", "高", "Chrome", "API返回空数组，前端显示el-empty", "待测"],
        ["H-022", "API联调", "加载状态 Skeleton 显示", "刷新页面时，确认表格先显示Skeleton，API响应后显示数据", "手动", "中", "Chrome", "Skeleton显示，数据加载后消失", "待测"],
        ["H-023", "API联调", "后端500错误处理", "模拟后端返回500，确认前端表格显示错误提示", "手动", "高", "Chrome", "显示错误状态，不显示空白", "待测"],
        ["H-024", "API联调", "网络断开时加载状态", "断网后筛选/分页，确认加载状态持续，网络恢复后刷新", "手动", "中", "Chrome", "加载状态持续，恢复后数据刷新", "待测"],
    ]
    for row in history_cases:
        ws.append(row)
    
    # ===== 4. Configuration =====
    ws = wb.create_sheet("Configuration")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    config_cases = [
        ["C-001", "模式", "只读模式", "默认进入，确认所有字段不可编辑", "手动", "高", "Chrome/Safari", "表单禁用，Edit按钮可见", "PASS/待测"],
        ["C-002", "模式", "编辑模式", "点击Edit按钮，确认字段变为可编辑", "手动", "高", "Chrome/Safari", "输入框可编辑，Save/Cancel按钮出现", "PASS/待测"],
        ["C-003", "模式", "取消编辑", "点击Cancel，确认恢复只读模式，修改不保存", "手动", "高", "Chrome/Safari", "表单恢复禁用，数据回滚", "PASS/待测"],
        ["C-004", "敏感字段", "API Token掩码", "只读模式下，确认API Token显示为掩码", "手动", "高", "Chrome/Safari", "掩码显示，眼睛图标可点击", "PASS/待测"],
        ["C-005", "敏感字段", "Token显示", "点击眼睛图标，确认Token显示5秒", "手动", "高", "Chrome/Safari", "Token明文显示，5秒后自动隐藏", "PASS/待测"],
        ["C-006", "敏感字段", "Webhook Secret掩码", "确认Webhook Secret同样掩码显示", "手动", "高", "Chrome/Safari", "掩码显示，可点击揭示", "PASS/待测"],
        ["C-007", "表单", "GitLab URL", "编辑模式下输入URL，确认格式验证", "手动", "高", "Chrome/Safari", "URL格式验证，错误提示红色", "PASS/待测"],
        ["C-008", "表单", "LLM Provider", "选择不同provider，确认模型选项变化", "手动", "高", "Chrome/Safari", "模型下拉框内容变化", "PASS/待测"],
        ["C-009", "表单", "Temperature滑块", "拖动滑块，确认数值实时变化", "手动", "中", "Chrome/Safari", "滑块移动，数值同步更新", "PASS/待测"],
        ["C-010", "表单", "保存按钮", "无修改时，确认Save按钮禁用", "手动", "高", "Chrome/Safari", "Save按钮灰色不可点", "PASS/待测"],
        ["C-011", "表单", "保存修改", "修改后点击Save，确认保存成功", "手动", "高", "Chrome/Safari", "Toast成功通知，Dirty状态清除", "PASS/待测"],
        ["C-012", "表单", "未保存警告", "修改后尝试离开页面，确认警告对话框", "手动", "高", "Chrome/Safari", "确认对话框弹出", "PASS/待测"],
        ["C-013", "表单", "加载状态", "刷新页面，确认先显示骨架屏", "手动", "中", "Chrome/Safari", "el-skeleton显示，加载后消失", "PASS/待测"],
        ["C-014", "表单", "错误状态", "输入错误URL，确认表单验证错误，Save禁用", "手动", "中", "Chrome/Safari", "红色错误提示，Save按钮禁用", "PASS/待测"],
        ["C-015", "连接测试", "LLM测试", "点击Test Connection，确认测试流程", "手动", "中", "Chrome/Safari", "按钮loading，结果显示latency", "PASS/待测"],
        ["C-016", "响应式", "移动端表单", "375px窗口，确认表单单列布局", "手动", "中", "Chrome", "表单项垂直堆叠", "PASS/待测"],
        ["C-017", "API联调", "GET /api/v1/config 加载完整配置", "curl 调用API，确认返回所有配置字段", "自动", "高", "所有", "JSON包含所有配置项，数据类型正确", "待测"],
        ["C-018", "API联调", "PUT /api/v1/config 保存同步到 AppConfig", "修改配置后保存，curl API确认PUT请求成功，数据持久化", "自动", "高", "Chrome", "PUT 200成功，再次GET确认数据已更新", "待测"],
        ["C-019", "API联调", "POST /api/v1/config/test 返回 latencyMs", "点击Test Connection，curl 对应API确认返回latencyMs", "自动", "高", "Chrome", "POST返回{latencyMs: number}，前端显示正确", "待测"],
        ["C-020", "API联调", "只读/编辑模式 API 数据一致性", "进入编辑模式后，确认表单初始值与GET /api/v1/config一致", "自动", "高", "Chrome", "表单初始值与API返回完全一致", "待测"],
        ["C-021", "API联调", "网络断开时保存失败处理", "编辑模式下断网，点击Save，确认显示网络错误提示", "手动", "高", "Chrome", "显示网络错误Toast，数据不丢失", "待测"],
        ["C-022", "API联调", "500错误时错误提示", "模拟后端返回500，确认前端显示错误提示，保存按钮恢复可用", "手动", "高", "Chrome", "显示错误Toast，按钮状态恢复", "待测"],
        ["C-023", "API联调", "useConfig composable 加载状态", "刷新页面时，确认useConfig加载状态触发Skeleton显示", "手动", "中", "Chrome", "加载状态为true时Skeleton显示，false时消失", "待测"],
        ["C-024", "API联调", "配置保存后全局状态同步", "保存配置后，切换至其他页面再返回，确认配置已持久化", "手动", "高", "Chrome", "返回配置页显示最新保存值", "待测"],
        ["C-025", "API联调", "GET/PUT 配置往返数据一致性", "GET获取配置→修改→PUT保存→再次GET，确认数据一致", "自动", "高", "所有", "PUT前后GET数据一致，无字段丢失", "待测"],
    ]
    for row in config_cases:
        ws.append(row)
    
    # ===== 5. Queue Monitor =====
    ws = wb.create_sheet("Queue Monitor")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    queue_cases = [
        ["Q-001", "统计", "活跃任务数", "确认统计行显示正确active数量", "手动", "高", "Chrome/Safari", "数字与任务卡片一致", "PASS/待测"],
        ["Q-002", "统计", "队列深度", "确认队列深度 = active + queued", "手动", "高", "Chrome/Safari", "计算正确", "PASS/待测"],
        ["Q-003", "统计", "进度条", "确认进度条显示正确百分比", "手动", "中", "Chrome/Safari", "颜色对应状态", "PASS/待测"],
        ["Q-004", "任务卡片", "状态显示", "确认running任务有进度条和专家名", "手动", "高", "Chrome/Safari", "卡片内容完整正确", "PASS/待测"],
        ["Q-005", "任务卡片", "取消操作", "点击Cancel按钮，确认有确认对话框", "手动", "中", "Chrome/Safari", "ElMessageBox确认", "PASS/待测"],
        ["Q-006", "任务卡片", "重试操作", "点击Retry按钮，确认任务移到queued", "手动", "中", "Chrome/Safari", "状态变化，Toast通知", "PASS/待测"],
        ["Q-007", "实时更新", "SSE进度", "等待2秒，确认进度条自动增加", "手动", "高", "Chrome/Safari", "进度条数值变化", "PASS/待测"],
        ["Q-008", "实时更新", "任务完成", "等待任务完成，确认自动消失", "手动", "中", "Chrome/Safari", "任务从active移除", "PASS/待测"],
        ["Q-009", "暂停", "暂停队列", "点击Pause按钮，确认队列暂停", "手动", "高", "Chrome/Safari", "按钮变Resume，queued显示暂停", "PASS/待测"],
        ["Q-010", "暂停", "恢复队列", "点击Resume按钮，确认恢复处理", "手动", "高", "Chrome/Safari", "按钮变Pause，任务继续", "PASS/待测"],
        ["Q-011", "空状态", "无任务", "清空所有任务，确认显示el-empty", "手动", "中", "Chrome/Safari", "空状态图标和文字", "PASS/待测"],
        ["Q-012", "加载", "骨架屏", "刷新页面，确认先显示骨架屏", "手动", "中", "Chrome/Safari", "el-skeleton显示", "PASS/待测"],
        ["Q-013", "API联调", "GET /api/v1/queue/stats 实时统计", "curl 调用API，确认返回active/queued/completed/failed/paused/isPaused", "自动", "高", "所有", "JSON包含所有统计字段，数值正确", "待测"],
        ["Q-014", "API联调", "POST /api/v1/queue/pause 暂停成功", "点击Pause按钮，curl API确认POST成功，stats.isPaused=true", "自动", "高", "Chrome", "POST 200，isPaused变为true，按钮变Resume", "待测"],
        ["Q-015", "API联调", "POST /api/v1/queue/resume 恢复成功", "点击Resume按钮，curl API确认POST成功，stats.isPaused=false", "自动", "高", "Chrome", "POST 200，isPaused变为false，按钮变Pause", "待测"],
        ["Q-016", "API联调", "POST /api/v1/queue/max-concurrent 调整并发", "修改最大并发数，curl API确认POST成功，新任务按新并发执行", "自动", "高", "Chrome", "POST 200，并发数调整生效", "待测"],
        ["Q-017", "API联调", "DELETE /api/v1/queue/tasks/{id} 取消任务", "点击Cancel并确认，curl API确认DELETE成功，任务从列表移除", "自动", "高", "Chrome", "DELETE 200，任务状态变为cancelled", "待测"],
        ["Q-018", "API联调", "isPaused 计算属性绑定正确", "通过API暂停/恢复队列，确认前端isPaused与API返回值一致", "自动", "高", "Chrome", "isPaused与API stats.isPaused始终一致", "待测"],
        ["Q-019", "API联调", "网络断开时状态显示", "断网后刷新页面，确认队列状态显示错误/不可用", "手动", "中", "Chrome", "显示网络错误，统计数字为—或错误提示", "待测"],
        ["Q-020", "API联调", "后端500错误处理", "模拟后端返回500，确认前端显示错误提示，不崩溃", "手动", "高", "Chrome", "显示错误状态，页面正常可用", "待测"],
    ]
    for row in queue_cases:
        ws.append(row)
    
    # ===== 6. LLM Status =====
    ws = wb.create_sheet("LLM Status")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    llm_cases = [
        ["L-001", "卡片", "提供商信息", "确认每个provider显示名称、状态、延迟", "手动", "高", "Chrome/Safari", "卡片内容完整", "PASS/待测"],
        ["L-002", "卡片", "状态颜色", "确认绿色=success、黄色=warning、红色=error", "手动", "高", "Chrome/Safari", "颜色对应正确", "PASS/待测"],
        ["L-003", "卡片", "延迟颜色", "确认延迟>1s为黄色、>2s为红色", "手动", "中", "Chrome/Safari", "颜色编码正确", "PASS/待测"],
        ["L-004", "卡片", "sparkline", "确认24h延迟趋势图显示", "手动", "中", "Chrome/Safari", "SVG折线图存在", "PASS/待测"],
        ["L-005", "测试", "单provider测试", "点击Test按钮，确认测试流程", "手动", "高", "Chrome/Safari", "按钮loading，结果显示", "PASS/待测"],
        ["L-006", "测试", "批量刷新", "点击Refresh All，确认所有provider刷新", "手动", "中", "Chrome/Safari", "所有卡片loading，Toast汇总", "PASS/待测"],
        ["L-007", "统计", "汇总行", "确认顶部汇总卡片数字正确", "手动", "中", "Chrome/Safari", "健康/降级/错误/离线数量正确", "PASS/待测"],
        ["L-008", "配置", "跳转配置", "点击Configure，确认跳转至Config页面", "手动", "中", "Chrome/Safari", "路由跳转至/#/config", "PASS/待测"],
        ["L-009", "加载", "骨架屏", "刷新页面，确认先显示骨架屏", "手动", "中", "Chrome/Safari", "el-skeleton显示", "PASS/待测"],
        ["L-010", "API联调", "GET /api/v1/llm/providers 返回真实数据", "curl 调用API，确认返回providers数组，包含名称/状态/延迟", "自动", "高", "所有", "JSON数组，每个provider有id/name/status/latency字段", "待测"],
        ["L-011", "API联调", "POST /api/v1/llm/providers/{id}/test 测试单个", "点击Test按钮，curl API确认POST成功，返回latencyMs和status", "自动", "高", "Chrome", "POST 200，返回{latencyMs, status}，前端卡片更新", "待测"],
        ["L-012", "API联调", "useLlmStatus composable 加载状态", "刷新页面时，确认加载状态触发Skeleton显示，API响应后消失", "手动", "中", "Chrome", "加载时Skeleton显示，加载完成后卡片渲染", "待测"],
        ["L-013", "API联调", "后端数据与前端卡片一致", "curl API获取数据，与前端卡片显示对比，确认一致", "自动", "高", "Chrome", "前端显示与API数据完全一致，无错位", "待测"],
        ["L-014", "API联调", "网络错误时重试机制", "断网后恢复，确认前端自动重新获取providers数据", "手动", "中", "Chrome", "网络恢复后自动刷新，无需手动操作", "待测"],
        ["L-015", "API联调", "500错误时错误提示", "模拟后端返回500，确认前端显示错误提示，不显示空白", "手动", "高", "Chrome", "显示错误状态/Toast，保留上次数据", "待测"],
        ["L-016", "API联调", "刷新按钮 API 调用正确", "点击Refresh All，通过Network面板确认发送正确GET请求", "手动", "高", "Chrome", "Network中GET /api/v1/llm/providers请求成功", "待测"],
        ["L-017", "API联调", "提供商延迟排序正确", "确认卡片按延迟从小到大或按状态优先级排序", "手动", "中", "Chrome", "排序逻辑与后端一致或符合预期", "待测"],
        ["L-018", "API联调", "空提供商列表处理", "后端返回空数组时，确认前端显示空状态提示", "手动", "中", "Chrome", "显示el-empty或提示配置LLM", "待测"],
    ]
    for row in llm_cases:
        ws.append(row)
    
    # ===== 7. System Logs =====
    ws = wb.create_sheet("System Logs")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    logs_cases = [
        ["LG-001", "显示", "终端风格", "确认日志背景为黑色，字体等宽", "手动", "高", "Chrome/Safari", "黑色背景，monospace字体", "PASS/待测"],
        ["LG-002", "显示", "级别颜色", "确认INFO白色、WARN黄色、ERROR红色、DEBUG灰色", "手动", "高", "Chrome/Safari", "颜色对应正确", "PASS/待测"],
        ["LG-003", "显示", "时间戳", "确认时间戳格式正确", "手动", "中", "Chrome/Safari", "时间戳可读", "PASS/待测"],
        ["LG-004", "过滤", "级别筛选", "取消勾选INFO，确认只显示其他级别", "手动", "高", "Chrome/Safari", "INFO日志消失", "PASS/待测"],
        ["LG-005", "过滤", "关键词搜索", "输入关键词，确认日志过滤", "手动", "高", "Chrome/Safari", "只显示匹配的日志行", "PASS/待测"],
        ["LG-006", "过滤", "清空搜索", "清除搜索框，确认恢复全部日志", "手动", "中", "Chrome/Safari", "所有日志恢复显示", "PASS/待测"],
        ["LG-007", "控制", "暂停", "点击Pause按钮，确认日志停止滚动", "手动", "高", "Chrome/Safari", "新日志不显示，计数增加", "PASS/待测"],
        ["LG-008", "控制", "恢复", "点击Resume按钮，确认日志继续", "手动", "高", "Chrome/Safari", "日志继续滚动，计数清零", "PASS/待测"],
        ["LG-009", "控制", "下载", "点击Download按钮，确认导出文件", "手动", "中", "Chrome/Safari", "NDJSON文件下载，Toast通知", "PASS/待测"],
        ["LG-010", "控制", "清除", "点击Clear按钮，确认日志清空", "手动", "中", "Chrome/Safari", "日志清空，确认对话框", "PASS/待测"],
        ["LG-011", "实时", "新日志提示", "暂停时等待新日志，确认浮动按钮出现", "手动", "中", "Chrome/Safari", "浮动按钮显示新日志数量", "PASS/待测"],
        ["LG-012", "实时", "自动滚动", "确认日志自动滚动到底部", "手动", "高", "Chrome/Safari", "最新日志始终在底部可见", "PASS/待测"],
        ["LG-013", "空状态", "清除后", "清除后确认显示等待日志空状态", "手动", "中", "Chrome/Safari", "el-empty显示", "PASS/待测"],
        ["LG-014", "响应式", "移动端", "375px窗口，确认字体缩小，按钮堆叠", "手动", "中", "Chrome", "布局适配", "PASS/待测"],
        ["LG-015", "API联调", "GET /api/v1/logs SSE 连接建立", "通过EventSource连接，确认HTTP 200及SSE响应头", "自动", "高", "所有", "Response headers包含Content-Type: text/event-stream", "待测"],
        ["LG-016", "API联调", "SSE 实时日志流接收", "等待后端产生日志，确认SSE流实时推送数据", "自动", "高", "Chrome", "EventSource onmessage触发，日志实时追加", "待测"],
        ["LG-017", "API联调", "GET /api/v1/logs/download NDJSON 下载", "点击Download，curl对应API确认返回NDJSON格式", "自动", "高", "Chrome", "下载文件为NDJSON，每行一个JSON对象", "待测"],
        ["LG-018", "API联调", "日志持久化到 ~/.config/review-engine/logs.ndjson", "确认后端日志写入文件，下载内容与文件一致", "自动", "高", "所有", "文件存在，下载内容与磁盘文件一致", "待测"],
        ["LG-019", "API联调", "useLogs composable 暂停/恢复", "点击Pause/Resume，确认useLogs状态变化正确，SSE连接控制正确", "手动", "中", "Chrome", "暂停时EventSource关闭/停止处理，恢复时重新连接", "待测"],
        ["LG-020", "API联调", "网络断开时 SSE 重连", "断网后恢复，确认SSE自动重连并接收新日志", "手动", "高", "Chrome", "网络恢复后自动重连，日志继续推送", "待测"],
        ["LG-021", "API联调", "大量日志时前端性能", "模拟大量日志推送，确认前端不卡顿，滚动流畅", "手动", "中", "Chrome", "日志数量>1000时，滚动和过滤无卡顿", "待测"],
        ["LG-022", "API联调", "过滤与后端日志级别一致", "筛选不同级别，确认前端过滤结果与后端日志级别匹配", "自动", "高", "Chrome", "INFO/WARN/ERROR/DEBUG筛选结果准确", "待测"],
    ]
    for row in logs_cases:
        ws.append(row)
    
    # ===== 8. Experts Management =====
    ws = wb.create_sheet("Experts Management")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    experts_cases = [
        ["E-001", "卡片", "显示信息", "确认每个卡片显示图标、名称、类别、权重、描述", "手动", "高", "Chrome/Safari", "卡片内容完整正确", "PASS/待测"],
        ["E-002", "卡片", "类别颜色", "确认不同类别有不同颜色标签", "手动", "中", "Chrome/Safari", "颜色区分不同类别", "PASS/待测"],
        ["E-003", "卡片", "启用禁用", "点击toggle开关，确认状态变化", "手动", "高", "Chrome/Safari", "开关状态变化，Toast通知", "PASS/待测"],
        ["E-004", "卡片", "禁用样式", "禁用后确认卡片变为灰色半透明", "手动", "中", "Chrome/Safari", "opacity降低，grayscale滤镜", "PASS/待测"],
        ["E-005", "筛选", "搜索", "输入专家名称，确认过滤", "手动", "高", "Chrome/Safari", "只显示匹配卡片", "PASS/待测"],
        ["E-006", "筛选", "类别筛选", "选择类别，确认只显示该类别", "手动", "中", "Chrome/Safari", "卡片按类别过滤", "PASS/待测"],
        ["E-007", "详情", "弹窗打开", "点击View Details，确认弹窗打开", "手动", "高", "Chrome/Safari", "ElDialog显示，内容完整", "PASS/待测"],
        ["E-008", "详情", "提示词预览", "确认弹窗显示专家提示词（monospace）", "手动", "中", "Chrome/Safari", "提示词正确显示，可复制", "PASS/待测"],
        ["E-009", "详情", "最近审查", "确认弹窗显示最近5条审查记录", "手动", "中", "Chrome/Safari", "表格5行，可点击", "PASS/待测"],
        ["E-010", "编辑", "编辑弹窗", "点击Edit按钮，确认编辑弹窗打开", "手动", "高", "Chrome/Safari", "表单可编辑", "PASS/待测"],
        ["E-011", "编辑", "权重滑块", "拖动权重滑块，确认数值变化", "手动", "中", "Chrome/Safari", "滑块移动，数值更新", "PASS/待测"],
        ["E-012", "编辑", "保存", "修改后保存，确认更新成功", "手动", "高", "Chrome/Safari", "Toast成功，卡片更新", "PASS/待测"],
        ["E-013", "统计", "顶部统计", "确认顶部显示active数和平均权重", "手动", "中", "Chrome/Safari", "数字正确", "PASS/待测"],
        ["E-014", "响应式", "网格布局", "确认1280px 4列、1024px 3列、768px 2列、375px 1列", "手动", "中", "Chrome", "断点正确", "PASS/待测"],
        ["E-015", "加载", "骨架屏", "刷新页面，确认先显示6个骨架卡片", "手动", "中", "Chrome/Safari", "el-skeleton显示", "PASS/待测"],
        ["E-016", "API联调", "GET /api/v1/system/experts 获取真实列表", "curl 调用API，确认返回experts数组，字段完整", "自动", "高", "所有", "JSON数组，每个expert有id/name/category/weight/enabled/description字段", "待测"],
        ["E-017", "API联调", "PUT /api/v1/system/experts/{id} 更新 enabled", "点击toggle开关，curl API确认PUT成功，enabled字段更新", "自动", "高", "Chrome", "PUT 200，enabled值切换，卡片状态更新", "待测"],
        ["E-018", "API联调", "PUT /api/v1/system/experts/{id} 更新 weight", "修改权重并保存，curl API确认PUT成功，weight字段更新", "自动", "高", "Chrome", "PUT 200，weight值更新，卡片显示新权重", "待测"],
        ["E-019", "API联调", "useExperts composable 加载状态", "刷新页面时，确认加载状态触发Skeleton，API响应后消失", "手动", "中", "Chrome", "Skeleton显示，加载完成后卡片渲染", "待测"],
        ["E-020", "API联调", "后端数据与前端卡片一致", "curl API获取数据，与前端卡片显示对比，确认一致", "自动", "高", "Chrome", "前端显示与API数据完全一致，无错位", "待测"],
        ["E-021", "API联调", "网络断开时保存失败", "编辑模式下断网，点击Save，确认显示网络错误", "手动", "中", "Chrome", "显示网络错误Toast，修改不丢失", "待测"],
        ["E-022", "API联调", "500错误时错误提示", "模拟后端返回500，确认前端显示错误提示", "手动", "高", "Chrome", "显示错误Toast，卡片状态不变", "待测"],
        ["E-023", "API联调", "专家权重后端验证", "PUT权重为非法值（如负数或>100），确认后端返回400错误", "自动", "高", "所有", "PUT 400，前端显示验证错误", "待测"],
        ["E-024", "API联调", "批量启用/禁用", "快速切换多个专家开关，确认每个PUT请求独立成功，无竞态", "自动", "高", "Chrome", "所有PUT成功，最终状态与前端一致", "待测"],
    ]
    for row in experts_cases:
        ws.append(row)
    
    # ===== 9. API Integration (新增) =====
    ws = wb.create_sheet("API Integration")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    api_cases = [
        ["AI-001", "数据一致性", "GET/PUT 配置往返验证", "GET获取配置→修改→PUT保存→再次GET，确认数据一致", "自动", "高", "所有", "PUT前后GET数据一致，无字段丢失或类型变化", "待测"],
        ["AI-002", "数据一致性", "Dashboard KPI 数据往返验证", "curl GET /api/v1/dashboard，与前端KPI显示对比，确认数值一致", "自动", "高", "Chrome", "前端KPI数值与API返回一致，无精度丢失", "待测"],
        ["AI-003", "数据一致性", "Review History 筛选结果一致性", "curl 带筛选参数调用API，与前端筛选结果对比记录数和内容", "自动", "高", "Chrome", "前端显示记录数与API返回总数一致", "待测"],
        ["AI-004", "数据一致性", "Experts 状态往返验证", "GET experts→修改enabled/weight→PUT→再次GET，确认状态持久化", "自动", "高", "Chrome", "修改后状态持久化，重启后仍一致", "待测"],
        ["AI-005", "错误处理", "网络断开时全局错误处理", "断网后操作各页面，确认统一错误提示，不崩溃", "手动", "高", "Chrome", "显示网络错误Toast/提示，页面可用", "待测"],
        ["AI-006", "错误处理", "后端500错误时 Toast 提示", "模拟后端返回500，确认各页面显示正确错误提示", "手动", "高", "Chrome", "显示500错误Toast，保留上次数据", "待测"],
        ["AI-007", "错误处理", "后端超时处理", "模拟后端延迟>10s，确认前端请求超时并显示提示", "手动", "高", "Chrome", "显示超时提示，取消loading状态", "待测"],
        ["AI-008", "错误处理", "认证/授权失败处理", "模拟401/403响应，确认前端跳转登录或显示权限提示", "手动", "中", "Chrome", "显示认证失败提示，不泄露敏感信息", "待测"],
        ["AI-009", "加载状态", "Skeleton 加载动画所有页面", "刷新各页面，确认均显示Skeleton加载状态", "手动", "高", "Chrome", "Dashboard/History/Config/Queue/LLM/Experts均有Skeleton", "待测"],
        ["AI-010", "加载状态", "加载状态与API响应同步", "确认Skeleton显示时间与API响应时间一致，响应后立即消失", "手动", "中", "Chrome", "API响应后Skeleton立即消失，无闪烁", "待测"],
        ["AI-011", "实时更新", "SSE 连接稳定性", "保持Logs页面长时间打开，确认SSE连接稳定不中断", "自动", "高", "Chrome", "SSE连接持续>30分钟不中断，自动重连正常", "待测"],
        ["AI-012", "实时更新", "Dashboard 轮询刷新数据", "等待60s自动刷新，确认Network面板显示新的GET请求", "手动", "高", "Chrome", "每60s发送一次GET /api/v1/dashboard，数据更新", "待测"],
        ["AI-013", "实时更新", "Queue Monitor 实时统计更新", "保持Queue页面打开，确认stats自动刷新，任务状态实时变化", "手动", "高", "Chrome", "stats自动更新，任务进度实时变化", "待测"],
        ["AI-014", "实时更新", "跨页面数据一致性", "在Dashboard修改数据后，切换至History/Queue，确认数据同步", "手动", "高", "Chrome", "各页面数据一致，无缓存滞后", "待测"],
        ["AI-015", "并发", "并发API请求处理", "快速切换页面或操作，确认并发请求不冲突，结果正确", "手动", "中", "Chrome", "并发请求无竞态，最终状态正确", "待测"],
        ["AI-016", "并发", "批量操作原子性", "批量操作（如批量禁用专家）时，确认每个请求独立处理，无丢失", "自动", "高", "Chrome", "所有请求成功，无丢失或重复", "待测"],
        ["AI-017", "缓存", "缓存一致性", "修改配置后，确认前端缓存失效，下次获取最新数据", "手动", "中", "Chrome", "修改后缓存清除，再次获取为最新值", "待测"],
        ["AI-018", "日志", "请求/响应日志", "确认前端Network面板可查看所有API请求和响应", "手动", "低", "Chrome", "所有API请求在Network面板可见，响应可查看", "待测"],
        ["AI-019", "状态同步", "前端状态与后端状态同步", "刷新页面后，确认前端状态与后端数据库一致，无差异", "自动", "高", "Chrome", "刷新后状态与后端完全一致", "待测"],
        ["AI-020", "恢复", "错误恢复机制", "API错误后，确认前端可自动恢复（如重试、刷新），无需手动刷新", "手动", "中", "Chrome", "错误恢复后页面正常，数据正确", "待测"],
    ]
    for row in api_cases:
        ws.append(row)
    
    # 设置所有 sheet 的列宽和样式
    for sheet in wb.worksheets:
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 55
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 8
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 40
        sheet.column_dimensions['I'].width = 12
        
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell_align
                cell.border = thin_border
    
    output_path = '/Users/isletspace/Workspace/gitlab.islet.space/review-engine/frontend-test-cases-v2.xlsx'
    wb.save(output_path)
    print(f"Saved to: {output_path}")
    
    # 统计
    for sheet in wb.worksheets:
        print(f"Sheet: {sheet.title}, Rows: {sheet.max_row - 1}")
    
    total = sum(s.max_row - 1 for s in wb.worksheets)
    print(f"Total cases: {total}")
    return output_path

if __name__ == '__main__':
    create_v2_workbook()
